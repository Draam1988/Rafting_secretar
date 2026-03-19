from __future__ import annotations
from pathlib import Path

from raftsecretary.domain.slalom import best_run_for_team
from raftsecretary.storage.competition_storage import load_competition_settings
from raftsecretary.storage.judges_storage import load_judges
from raftsecretary.storage.slalom_storage import load_slalom_runs, load_slalom_lineup_flags
from raftsecretary.storage.team_storage import load_teams
from raftsecretary.export.pdf_builder import (
    new_pdf, pdf_bytes, write_doc_header, write_category_header,
    write_table_header, write_footer, CONTENT_W, MARGIN,
)
from raftsecretary.export.xlsx_builder import (
    new_workbook, workbook_bytes, write_meta_row, write_header_row, BORDER,
)
from raftsecretary.export.sprint_export import _fmt, _judge_name, _category_label, _lineup_text
from raftsecretary.web.app import _slalom_places_map, _slalom_scored_run


def _fmt_hhmmss(total_seconds: int) -> str:
    if not total_seconds:
        return ""
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def _load_data(db_path: Path):
    settings = load_competition_settings(db_path)
    judges = load_judges(db_path)
    teams = load_teams(db_path)
    comp_dates = ", ".join(settings.competition_dates or [settings.competition_date])
    return settings, judges, teams, comp_dates


def build_slalom_pdf(db_path: Path) -> bytes:
    settings, judges, teams, comp_dates = _load_data(db_path)
    gate_count = settings.slalom_gate_count or 8
    pdf = new_pdf()
    pdf.add_page()
    write_doc_header(pdf, "Слалом", settings.name or "", comp_dates,
                     settings.venue or "", settings.organizer or "")

    # Рассчитываем ширину ворот
    fixed_w = 8 + 30 + 25 + 35 + 18 + 15 + 15 + 15 + 10  # сумма фиксированных колонок
    gate_w = max(5.0, min(8.0, (CONTENT_W - fixed_w) / max(gate_count, 1)))

    fixed_cols = [
        ("№", 8), ("Команда", 30), ("Субъект", 25), ("Состав", 35),
        ("Попытка", 18), ("Старт", 15), ("Финиш", 15),
    ]
    gate_cols = [(str(i), gate_w) for i in range(1, gate_count + 1)]
    tail_cols = [("Итог", 15), ("Место", 10)]
    all_cols = fixed_cols + gate_cols + tail_cols

    for category in settings.categories:
        category_teams = [t for t in teams if t.category_key == category.key]
        slalom_runs = load_slalom_runs(db_path, category.key)
        if not category_teams and not slalom_runs:
            continue

        slalom_places = _slalom_places_map(slalom_runs)
        grouped_runs: dict = {}
        scored_by_team: dict = {}
        for run in slalom_runs:
            grouped_runs.setdefault(run.team_name, []).append(run)
            scored = _slalom_scored_run(run)
            if scored is not None:
                scored_by_team.setdefault(run.team_name, []).append(scored)
        lineup_flags = load_slalom_lineup_flags(db_path, category.key)

        write_category_header(pdf, _category_label(category))
        write_table_header(pdf, all_cols)

        for team in sorted(category_teams, key=lambda t: (slalom_places.get(t.name, 9999), t.start_number)):
            runs = grouped_runs.get(team.name, [])
            scored_runs = scored_by_team.get(team.name, [])
            best = best_run_for_team(scored_runs) if scored_runs else None
            place = slalom_places.get(team.name)
            place_text = str(place) if place else ""
            run_by_attempt = {getattr(r, "attempt_number", 0): r for r in runs}
            scored_by_att = {r.attempt_number: r for r in scored_runs}
            lineup = _lineup_text(team, lineup_flags)

            for attempt_num in (1, 2):
                raw = run_by_attempt.get(attempt_num)
                scored = scored_by_att.get(attempt_num)
                is_best = best is not None and best.attempt_number == attempt_num
                attempt_label = f"{attempt_num}-я" + (" *" if is_best else "")
                start_text = finish_text = total_text = ""
                if raw is not None:
                    base_s = int(getattr(raw, "base_time_seconds", 0))
                    fin_s = int(getattr(raw, "finish_time_seconds", 0))
                    if base_s > 0:
                        start_text = _fmt_hhmmss(base_s)
                    if fin_s > 0:
                        finish_text = _fmt_hhmmss(fin_s)
                if scored is not None:
                    total_text = _fmt(scored.total_time_seconds)
                gate_penalties = list(getattr(raw, "gate_penalties", [])) if raw else []
                if len(gate_penalties) < gate_count:
                    gate_penalties += [0] * (gate_count - len(gate_penalties))

                pdf.set_font("DejaVu", "", 8)
                if attempt_num == 1:
                    pdf.cell(8, 5, str(team.start_number), border=1)
                    pdf.cell(30, 5, team.name[:18], border=1)
                    pdf.cell(25, 5, team.region[:14], border=1)
                    pdf.cell(35, 5, lineup[:22], border=1)
                else:
                    pdf.cell(8, 5, "", border=1)
                    pdf.cell(30, 5, "", border=1)
                    pdf.cell(25, 5, "", border=1)
                    pdf.cell(35, 5, "", border=1)
                pdf.cell(18, 5, attempt_label, border=1)
                pdf.cell(15, 5, start_text, border=1, align="C")
                pdf.cell(15, 5, finish_text, border=1, align="C")
                for gp in gate_penalties[:gate_count]:
                    pdf.cell(gate_w, 5, str(gp) if gp else "", border=1, align="C")
                pdf.cell(15, 5, total_text, border=1, align="C")
                pdf.cell(10, 5, place_text if attempt_num == 1 else "", border=1, align="C")
                pdf.ln()
        pdf.ln(3)

    write_footer(pdf, _judge_name(judges.chief_judge), _judge_name(judges.chief_secretary))
    return pdf_bytes(pdf)


def build_slalom_xlsx(db_path: Path) -> bytes:
    from openpyxl.styles import Font as XFont, Alignment
    from raftsecretary.export.xlsx_builder import write_header_row
    settings, judges, teams, comp_dates = _load_data(db_path)
    gate_count = settings.slalom_gate_count or 8
    wb, ws = new_workbook()
    ws.title = "Слалом"

    row = 1
    write_meta_row(ws, row, "Соревнование", settings.name or ""); row += 1
    write_meta_row(ws, row, "Даты", comp_dates); row += 1
    write_meta_row(ws, row, "Место", settings.venue or ""); row += 1
    write_meta_row(ws, row, "Организатор", settings.organizer or ""); row += 1
    row += 1

    header_cols = [
        ("№", 6), ("Команда", 22), ("Субъект", 18), ("Состав", 30),
        ("Попытка", 12), ("Старт", 10), ("Финиш", 10),
    ] + [(str(i) + "в", 5) for i in range(1, gate_count + 1)] + [("Итог", 10), ("Место", 8)]

    for category in settings.categories:
        category_teams = [t for t in teams if t.category_key == category.key]
        slalom_runs = load_slalom_runs(db_path, category.key)
        if not category_teams and not slalom_runs:
            continue

        slalom_places = _slalom_places_map(slalom_runs)
        grouped_runs: dict = {}
        scored_by_team: dict = {}
        for run in slalom_runs:
            grouped_runs.setdefault(run.team_name, []).append(run)
            scored = _slalom_scored_run(run)
            if scored is not None:
                scored_by_team.setdefault(run.team_name, []).append(scored)
        lineup_flags = load_slalom_lineup_flags(db_path, category.key)

        ws.cell(row=row, column=1, value=_category_label(category)).font = XFont(bold=True, size=11)
        row += 1
        write_header_row(ws, row, header_cols); row += 1

        for team in sorted(category_teams, key=lambda t: (slalom_places.get(t.name, 9999), t.start_number)):
            runs = grouped_runs.get(team.name, [])
            scored_runs = scored_by_team.get(team.name, [])
            best = best_run_for_team(scored_runs) if scored_runs else None
            place = slalom_places.get(team.name)
            run_by_attempt = {getattr(r, "attempt_number", 0): r for r in runs}
            scored_by_att = {r.attempt_number: r for r in scored_runs}
            lineup = _lineup_text(team, lineup_flags)

            for attempt_num in (1, 2):
                raw = run_by_attempt.get(attempt_num)
                scored = scored_by_att.get(attempt_num)
                is_best = best is not None and best.attempt_number == attempt_num
                attempt_label = f"{attempt_num}-я попытка" + (" (лучшая)" if is_best else "")
                start_text = finish_text = total_text = ""
                if raw is not None:
                    base_s = int(getattr(raw, "base_time_seconds", 0))
                    fin_s = int(getattr(raw, "finish_time_seconds", 0))
                    if base_s > 0:
                        start_text = _fmt_hhmmss(base_s)
                    if fin_s > 0:
                        finish_text = _fmt_hhmmss(fin_s)
                if scored is not None:
                    total_text = _fmt(scored.total_time_seconds)
                gate_penalties = list(getattr(raw, "gate_penalties", [])) if raw else []
                if len(gate_penalties) < gate_count:
                    gate_penalties += [0] * (gate_count - len(gate_penalties))

                col = 1
                if attempt_num == 1:
                    for val in [team.start_number, team.name, team.region, lineup]:
                        c = ws.cell(row=row, column=col, value=val)
                        c.border = BORDER; c.alignment = Alignment(vertical="top", wrap_text=True)
                        col += 1
                else:
                    for _ in range(4):
                        ws.cell(row=row, column=col).border = BORDER; col += 1
                for val in [attempt_label, start_text, finish_text]:
                    c = ws.cell(row=row, column=col, value=val)
                    c.border = BORDER; col += 1
                for gp in gate_penalties[:gate_count]:
                    c = ws.cell(row=row, column=col, value=gp if gp else "")
                    c.border = BORDER; col += 1
                ws.cell(row=row, column=col, value=total_text).border = BORDER; col += 1
                ws.cell(row=row, column=col, value=place if attempt_num == 1 else "").border = BORDER
                row += 1
        row += 1

    row += 1
    write_meta_row(ws, row, "Главный судья", _judge_name(judges.chief_judge)); row += 1
    write_meta_row(ws, row, "Главный секретарь", _judge_name(judges.chief_secretary))
    return workbook_bytes(wb)
