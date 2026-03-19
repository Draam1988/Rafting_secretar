from __future__ import annotations
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="D0D0D0")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def new_workbook() -> tuple[Workbook, object]:
    wb = Workbook()
    ws = wb.active
    return wb, ws


def write_meta_row(ws, row: int, label: str, value: str) -> None:
    ws.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws.cell(row=row, column=2, value=value)


def write_header_row(ws, row: int, cols: list[tuple[str, int]]) -> None:
    """cols: list of (label, width_chars)"""
    for col_idx, (label, width) in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_data_row(ws, row: int, values: list, col_count: int) -> None:
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value if value is not None else "")
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def workbook_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
